"""Tests de lógica pura (sin GCP): precios, filtros, SKU, fotos, totales, Excel, JWT."""
from __future__ import annotations

import datetime as dt
import io
import os
import sys
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ.setdefault("JWT_KEY", "test-key-no-usar")

import auth  # noqa: E402
import catalog  # noqa: E402
import fotos  # noqa: E402
import pedidos  # noqa: E402


def _df():
    rows = [
        # producto, nombre, marca, temporada, rubro, subrubro, color_cod, color, talle, sku, ean, stock, precio1, precio4
        ("M211", "Mochila Soft Rainbow", "Chimola", "DDN25", "Mochilas", None, "2059", "AQUA", "U", "M211_U_2059", "779", 99, 32900, 65800),
        ("M211", "Mochila Soft Rainbow", "Chimola", "DDN25", "Mochilas", None, "2058", "RAINBOW", "U", "M211_U_2058", "780", 107, 32900, 65800),
        ("VEST2", "Vestido Lino", "Lima", "AW26", "Vestidos", None, "2018", "GREEN", "8", "VEST2_8_2018", "", 3, 0, 50000),
        ("VEST2", "Vestido Lino", "Lima", "AW26", "Vestidos", None, "2018", "GREEN", "10", "VEST2_10_2018", "", 2, 0, 50000),
    ]
    cols = ["producto_cod", "producto_nombre", "marca", "temporada", "rubro", "subrubro", "color_cod", "color",
            "talle", "sku", "ean", "stock", "precio1", "precio4"]
    df = pd.DataFrame(rows, columns=cols)
    for c in catalog.PRECIO_COLS:
        if c not in df:
            df[c] = 0.0
    df["descvta"] = 0.0
    df["descripcion"] = ""
    df["categoria"] = ["Marroquineria", "Marroquineria", "Indumentaria", "Indumentaria"]
    return df


def test_precio_por_lista_y_sin_precio():
    df = catalog.con_precio(_df(), 1)
    assert df.loc[df.sku == "M211_U_2059", "precio"].iloc[0] == 32900
    assert pd.isna(df.loc[df.sku == "VEST2_8_2018", "precio"].iloc[0])  # lista 1 sin precio → no se vende
    df4 = catalog.con_precio(_df(), 4)
    assert df4.loc[df4.sku == "VEST2_8_2018", "precio"].iloc[0] == 50000
    assert catalog.columna_precio(0) == "precio1" and catalog.columna_precio(11) == "precio1"


def test_descuento_cabecera():
    assert catalog.aplicar_descuento(1000, 30) == 700
    assert catalog.aplicar_descuento(1000, 0) == 1000
    assert catalog.aplicar_descuento(32900, 20) == 26320


def test_filtros_y_busqueda():
    df = catalog.con_precio(_df(), 1)
    assert len(catalog.filtrar_variantes(df, {"marca": ["Lima"]})) == 2
    assert len(catalog.filtrar_variantes(df, {}, "m211 rainbow")) == 2  # el nombre también dice Rainbow
    assert len(catalog.filtrar_variantes(df, {}, "m211 aqua")) == 1
    assert len(catalog.filtrar_variantes(df, {}, "779")) == 1
    ops = catalog.opciones_filtros(df, {"marca": ["Lima"]})
    assert ops["temporada"] == ["AW26"] and ops["marca"] == ["Chimola", "Lima"]


def test_productos_agregado_y_detalle():
    df = catalog.con_precio(_df(), 1)
    p = catalog.productos(df)
    m = p[p.producto_cod == "M211"].iloc[0]
    assert m.stock == 206 and m.n_variantes == 2 and m.colores == ["AQUA", "RAINBOW"]
    d = catalog.get_producto(df, "VEST2")
    assert d["precio"] is None and d["talles"] == ["8", "10"]
    assert catalog.get_producto(df, "NOEXISTE") is None


def test_parseo_fotos():
    files = ["M211 AQUA (2).jpg", "M211 AQUA (1).jpg", "M211 LIGHT PINK (1).jpg", "M211 allcolors.jpg", "notas.txt.jpg"]
    out = fotos.parsear_fotos("M211", files, ["AQUA", "LIGHT PINK"])
    assert out[0]["filename"] == "M211 allcolors.jpg" and out[0]["is_cover"]
    aqua = fotos.fotos_por_color(out, "Aqua")
    assert [f["filename"] for f in aqua] == ["M211 AQUA (1).jpg", "M211 AQUA (2).jpg"]
    pink = fotos.fotos_por_color(out, "LIGHT PINK")
    assert [f["filename"] for f in pink] == ["M211 LIGHT PINK (1).jpg"]
    # sin paréntesis: "L11 VISON.jpg" → color VISON
    out2 = fotos.parsear_fotos("L11", ["L11 VISON.jpg", "L11 MINT (2).jpg"], ["VISON", "MINT"])
    assert {f["color_norm"] for f in out2} == {"VISON", "MINT"}
    assert fotos.parsear_fotos("X", []) == []
    assert fotos.norm("Light  Pink") == "LIGHTPINK" and fmt_ok()


def fmt_ok():
    return fotos.norm("Visón") == "VISON"


def test_carrito_y_totales():
    items = pedidos.agregar_al_carrito([], {"sku": "A", "cantidad": 2, "precio_unit": 100, "stock": 10})
    items = pedidos.agregar_al_carrito(items, {"sku": "A", "cantidad": 3, "precio_unit": 100, "stock": 10})
    items = pedidos.agregar_al_carrito(items, {"sku": "B", "cantidad": 1, "precio_unit": 50.5, "stock": 1})
    assert len(items) == 2 and items[0]["cantidad"] == 5
    tot = pedidos.calcular_totales(items, 30)
    assert tot["unidades"] == 6 and tot["subtotal"] == 550.5
    assert tot["total"] == round(550.5 * 0.7, 2) and tot["descuento_monto"] == round(550.5 - tot["total"], 2)


def _pedido():
    ahora = dt.datetime(2026, 8, 21, 15, 30, tzinfo=dt.timezone.utc)
    items = [
        {"sku": "M211_U_2059", "ean": "779", "producto_cod": "M211", "producto_nombre": "Mochila Soft Rainbow",
         "color_cod": "2059", "color": "AQUA", "talle": "U", "cantidad": 3, "precio_unit": 32900.0},
        {"sku": "M211_U_2058", "ean": "780", "producto_cod": "M211", "producto_nombre": "Mochila Soft Rainbow",
         "color_cod": "2058", "color": "RAINBOW", "talle": "U", "cantidad": 1, "precio_unit": 32900.0},
    ]
    tot = pedidos.calcular_totales(items, 20)
    p = {"numero": 12, "cliente_cod": 2722, "cliente_nombre": "Franquicia Jujuy", "cliente_cuit": "30-1",
         "usuario_email": "x@y.com", "lista_precios": 1, "items": items, **tot, "estado": "confirmado",
         "observaciones": "urgente", "confirmed_at": ahora, "fecha_str": "21/08/2026 12:30"}
    p["xlsx_filename"] = pedidos.nombre_archivo(p)
    return p


def test_excel_y_nombres():
    p = _pedido()
    assert p["xlsx_filename"] == "pedido_2722_12_20260821_123000.xlsx"
    assert pedidos.gcs_path(p) == "2026-08/pedido_2722_12_20260821_123000.xlsx"
    data = pedidos.generar_excel(p)
    assert data[:2] == b"PK" and len(data) > 2000
    import openpyxl  # viene con pandas? no siempre — si falta, solo validamos el zip
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Resumen", "Detalle"]
    det = wb["Detalle"]
    assert det["A2"].value == "M211" and det["H2"].value == 3 and det["I2"].value == 32900
    res = wb["Resumen"]
    vals = {res.cell(r, 1).value: res.cell(r, 2).value for r in range(1, 25)}
    assert vals["TOTAL"] == p["total"] == round(4 * 32900 * 0.8, 2)


def test_password_y_jwt():
    h = auth.hash_password("secreta123")
    assert auth.verify_password("secreta123", h) and not auth.verify_password("otra", h)
    assert not auth.verify_password("x", None)
    tok = auth.create_jwt({"email": "a@b.com", "cliente_cod": 2722, "rol": "cliente", "nombre_display": "A"})
    claims = auth.verify_jwt(tok)
    assert claims["sub"] == "a@b.com" and claims["cliente_cod"] == 2722
    assert auth.verify_jwt(tok + "x") is None and auth.verify_jwt(None) is None
    vencido = auth.create_jwt({"email": "a@b.com", "cliente_cod": 1}, ttl_hours=-1)
    assert auth.verify_jwt(vencido) is None


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-q"]))


# --- Fase 8: taxonomía y facetas nuevas ---
def test_normalizar_taxonomia():
    s = pd.Series(["Bolsos y Totes", "Bolsos y totes", "Librería", "Libreria", None, "  "])
    out = catalog.normalizar_taxonomia(s)
    assert out.iloc[0] == out.iloc[1]          # mayúsculas unificadas
    assert out.iloc[2] == out.iloc[3]          # acentos unificados
    assert out.iloc[4] == "Otros" and out.iloc[5] == "Otros"
    assert out.iloc[0][0].isupper()


def test_facetas_categoria_color_talle():
    df = catalog.con_precio(_df(), 1)
    ops = catalog.opciones_filtros(df, {})
    assert ops["categoria"] == ["Indumentaria", "Marroquineria"]
    assert ops["color"] == ["AQUA", "GREEN", "RAINBOW"]
    assert ops["talle"] == ["8", "10", "U"]    # orden natural de talles
    assert len(catalog.filtrar_variantes(df, {"categoria": ["Indumentaria"]})) == 2
    assert len(catalog.filtrar_variantes(df, {"color": ["AQUA"], "talle": ["U"]})) == 1
    assert catalog.productos(df).set_index("producto_cod").loc["M211", "categoria"] == "Marroquineria"


# --- Bugfix 2026-08-24: carrito por encima del stock + matriz de variantes ---
def test_agregar_al_carrito_capea_al_stock():
    items = pedidos.agregar_al_carrito([], {"sku": "A", "cantidad": 8, "precio_unit": 100, "stock": 10})
    items = pedidos.agregar_al_carrito(items, {"sku": "A", "cantidad": 5, "precio_unit": 100, "stock": 10})
    assert items[0]["cantidad"] == 10  # 8+5 pero hay 10
    # sin stock conocido no capea (compat con items viejos)
    items2 = pedidos.agregar_al_carrito([{"sku": "B", "cantidad": 3, "precio_unit": 1, "stock": None}],
                                        {"sku": "B", "cantidad": 4, "precio_unit": 1, "stock": None})
    assert items2[0]["cantidad"] == 7


def test_variantes_de_get_producto_sirven_para_item_de_carrito():
    import compra_rapida as cr
    df = catalog.con_precio(_df(), 1)
    prod = catalog.get_producto(df, "M211")
    v = prod["variantes"][0]
    it = cr.item_desde_variante(v, 3)   # no debe faltar ninguna key
    assert it["producto_cod"] == "M211" and it["producto_nombre"] == "Mochila Soft Rainbow"
    assert it["cantidad"] == 3 and it["sku"] == v["sku"] and it["manual"] is False


# --- Contacto por cliente + validación de email (2026-08-25) ---
def test_crear_usuario_valida_email():
    for malo in ("", "   ", "sinarroba", "con espacio@x.com", "con/barra@x.com", "x@sinpunto"):
        with pytest.raises(ValueError):
            auth.crear_usuario(malo, "pwd", None, "X")


def test_override_cliente_contacto_y_cuit(monkeypatch):
    import overrides
    ov = {1026: {"contacto_nombre": "Fernando", "contacto_email": "fer@kinderland.com.ar",
                 "cuit": "30-99999999-9"}}
    monkeypatch.setattr(overrides, "get_clientes_overrides", lambda: ov)
    e = overrides.aplicar_override_cliente({"cliente_cod": 1026, "descuento": 25.0,
                                            "lista_precios": 1, "cuit": "30-71065547-9"})
    assert e["contacto_nombre"] == "Fernando" and e["contacto_email"] == "fer@kinderland.com.ar"
    assert e["cuit"] == "30-99999999-9" and e["cuit_origen"] == "Override"
    monkeypatch.setattr(overrides, "get_clientes_overrides", lambda: {})
    e2 = overrides.aplicar_override_cliente({"cliente_cod": 1026, "descuento": 25.0,
                                             "lista_precios": 1, "cuit": "30-71065547-9"})
    assert e2["cuit"] == "30-71065547-9" and e2["cuit_origen"] == "Aleph"
    assert e2["contacto_nombre"] == "" and e2["contacto_email"] == ""


# --- Fase 9: foto ↔ variante (port del imagenes.csv del Woo) ---
def test_foto_por_color_y_fuzzy():
    files = ["M211 AQUA (1).jpg", "M211 AQUA (2).jpg", "M211 RAINBOW (1).jpg",
             "M211 allcolors.jpg", "M211 LIGTH PINK (1).jpg"]   # typo de LIGHT PINK
    mapa = fotos.foto_por_color("M211", ["AQUA", "RAINBOW", "LIGHT PINK"], files)
    assert mapa["AQUA"] == "M211 AQUA (1).jpg"
    assert mapa["RAINBOW"] == "M211 RAINBOW (1).jpg"
    assert mapa["LIGHTPINK"] == "M211 LIGTH PINK (1).jpg"   # fuzzy rescata el typo
    mapa2 = fotos.foto_por_color("M211", ["AQUA"], files, {"AQUA": "M211 AQUA (2).jpg"})
    assert mapa2["AQUA"] == "M211 AQUA (2).jpg"             # manual pisa al automático
    assert fotos._resolver_color_fuzzy("AZUL", ["AZULCLARO", "AZULOSCURO"]) is None  # ambiguo


def test_mapeo_variantes():
    variantes = [
        {"producto_cod": "M211", "sku": "M211_U_1", "color": "AQUA", "talle": "U"},
        {"producto_cod": "M211", "sku": "M211_U_2", "color": "VERDE", "talle": "U"},
        {"producto_cod": "ZZZ9", "sku": "ZZZ9_U_1", "color": "BLACK", "talle": "U"},
    ]
    idx = {"M211": ["M211 AQUA (1).jpg", "M211 allcolors.jpg"]}
    por_sku = {r["sku"]: r for r in fotos.mapeo_variantes(variantes, indice=idx)}
    assert por_sku["M211_U_1"]["origen"] == "color" and por_sku["M211_U_1"]["foto"] == "M211 AQUA (1).jpg"
    assert por_sku["M211_U_2"]["origen"] == "portada" and por_sku["M211_U_2"]["foto"] == "M211 allcolors.jpg"
    assert por_sku["ZZZ9_U_1"]["origen"] == "sin_foto"
    rows2 = fotos.mapeo_variantes(variantes, indice=idx,
                                  overrides_por_prod={"M211": {"fotos_color": {"VERDE": "M211 allcolors.jpg"}}})
    assert {r["sku"]: r for r in rows2}["M211_U_2"]["origen"] == "manual"


# --- Fase 10: import de historial de Aleph ---
def test_armar_pedido_import_aleph():
    import aleph_import
    cab = {"id": 10164, "tipo": 90, "numero": "A0004-00001565",
           "fecha": dt.date(2016, 4, 4), "pordscto": 20.0, "lis_pre": 1}
    items = [
        {"producto": "1569", "preuni": 340.0, "cantidad": 4, "talle": "U", "color_cod": 1997, "item": 1},
        {"producto": "1979", "preuni": 350.0, "cantidad": 2, "talle": "U", "color_cod": 1997, "item": 2},
    ]
    cliente = {"cliente_cod": 622, "cuit": "30714361984", "nombre_display": "DON BOSCO SRL"}
    p = aleph_import.armar_pedido(cab, items, {"1569": "CARTERA TRIANGULAR PREMIUM"},
                                  {1997: "UNICO"}, cliente, "bcaplan@lautin.com.ar", 99,
                                  ahora=dt.datetime(2026, 8, 26, tzinfo=dt.timezone.utc))
    assert p["np_aleph"] == "A0004-00001565" and p["estado"] == "procesado"
    assert p["subtotal"] == 340 * 4 + 350 * 2 == 2060
    assert p["descuento_pct"] == 20 and p["total"] == round(2060 * 0.8, 2)
    assert p["unidades"] == 6 and p["numero"] == 99
    assert p["items"][0]["sku"] == "1569_U_1997" and p["items"][0]["color"] == "UNICO"
    assert p["items"][0]["producto_nombre"] == "CARTERA TRIANGULAR PREMIUM"
    assert p["items"][1]["producto_nombre"] == "1979"   # sin nombre → código
    assert p["historial"][0]["por"] == "import-aleph"
    assert "NP A0004-00001565" in p["observaciones"] and p["fecha_str"] == "04/04/2016"


# --- Fase 11: reposición sugerida (franquicias) ---
def test_calcular_sugerido():
    import reposicion as rp
    assert rp.calcular_sugerido(2.0, 10, 21, 1, 999) == 32      # 42 − 10
    assert rp.calcular_sugerido(2.0, -5, 21, 1, 999) == 42      # stock negativo → 0
    assert rp.calcular_sugerido(2.0, 10, 21, 6, 999) == 36      # múltiplo U.B. hacia arriba
    assert rp.calcular_sugerido(2.0, 10, 21, 1, 20) == 20       # cap al disponible
    assert rp.calcular_sugerido(2.0, 10, 21, 6, 20) == 18       # cap respetando múltiplo
    assert rp.calcular_sugerido(0.1, 50, 21, 1, 999) == 0       # cubierto → 0
    assert rp.calcular_sugerido(2.0, 0, 21, 1, 0) == 0          # sin disponible → 0


def test_cruzar_con_catalogo():
    import reposicion as rp
    cat = catalog.con_precio(_df(), 1)
    vista = pd.DataFrame([
        {"ean": "779", "producto_cod": "X", "color": "X", "talle": "X",
         "stock_pv": 3, "vendidas_30d": 12, "vel_30d": 0.4, "cobertura_dias": 7.5},   # matchea por EAN
        {"ean": "", "producto_cod": "M211", "color": "RAINBOW", "talle": "U",
         "stock_pv": -2, "vendidas_30d": 6, "vel_30d": 0.2, "cobertura_dias": 2.0},   # matchea por prod/color/talle
        {"ean": "999999999", "producto_cod": "NOEX", "color": "Z", "talle": "U",
         "stock_pv": 1, "vendidas_30d": 1, "vel_30d": 0.03, "cobertura_dias": 30.0},  # no matchea
    ])
    out = rp.cruzar_con_catalogo(vista, cat)
    assert set(out["sku"]) == {"M211_U_2059", "M211_U_2058"}
    assert out.iloc[0]["sku"] == "M211_U_2058"   # cobertura 2.0 primero (urgente)
    assert int(out[out["sku"] == "M211_U_2059"].iloc[0]["stock_pv"]) == 3


def test_es_color_claro():
    from fotos import es_color_claro, color_claro
    for c in ["WHITE", "BLANCO", "OFF WHITE", "BEIGE", "NUDE", "PALE PINK", "LIGHT BLUE", "LT GREY",
              "CELESTE", "LILA", "MINT", "ROSA", "PINK AND GREEN", "PURPLE AND LIGHT BLUE", "Crudo"]:
        assert es_color_claro(c), c
    for c in ["BLACK", "NEGRO", "FULL BLACK", "CAMEL", "GRIS", "PURPLE", "BLUE", "DARK PINK",
              "NUDE & NEGRO", "BROWN PINK", "UNICO", "COBALT", "CLIMA", "", None]:
        assert not es_color_claro(c), c
    # preferencia: el más luminoso primero
    assert color_claro(["BLACK", "PALE PINK", "WHITE", "CELESTE"]) == "WHITE"
    assert color_claro(["BLACK", "CELESTE", "PALE PINK"]) == "PALE PINK"
    assert color_claro(["BLACK", "CAMEL"]) is None
    assert color_claro([]) is None


def test_foto_card_por_modo():
    from fotos import foto_card_filename, colores_por_modo, es_color_negro
    files = ["X1 BLACK (1).jpg", "X1 BLACK (2).jpg", "X1 PALE PINK (1).jpg", "X1 CAMEL (1).jpg"]
    cols = ["BLACK", "PALE PINK", "CAMEL"]
    assert foto_card_filename("X1", cols, "claro", files, {}) == ("X1 PALE PINK (1).jpg", True)
    assert foto_card_filename("X1", cols, "negro", files, {}) == ("X1 BLACK (1).jpg", True)
    # sin variante del tono → portada, match False
    assert foto_card_filename("X1", ["CAMEL"], "claro", files, {}) == ("X1 BLACK (1).jpg", False)
    # variante clara existe en catálogo pero SIN foto → no cuenta como match
    assert foto_card_filename("X1", ["BLACK", "WHITE"], "claro", files, {})[1] is False
    assert foto_card_filename("X1", cols, "claro", [], {}) == (None, False)
    assert es_color_negro("FULL BLACK") and es_color_negro("NEGRO") and not es_color_negro("BLUE")
    assert colores_por_modo(["NAVE BLACK", "BLACK", "WHITE"], "negro") == ["BLACK", "NAVE BLACK"]
    assert colores_por_modo(["CELESTE", "WHITE", "BLACK"], "claro") == ["WHITE", "CELESTE"]


def test_odoo_export_dos_hojas_sin_lineas_de_mas():
    import io
    import openpyxl
    import odoo_export as oe
    items = [
        {"producto_cod": "M269", "color": "CAMEL", "talle": "U", "cantidad": 2},
        {"producto_cod": "J23", "color": "AQUA", "talle": "6", "cantidad": 1},
        {"producto_cod": "J23", "color": "AQUA", "talle": "8", "cantidad": 1},
        {"producto_cod": "TX110", "color": "PINK", "talle": "U", "cantidad": 10},
        {"producto_cod": "X", "color": "Y", "talle": "U", "cantidad": 0},   # se ignora
    ]
    sin, indu = oe.armar_filas(items, "Drago Tech, Soporte Drago Leonel")
    assert sin == [["Drago Tech, Soporte Drago Leonel", "Comodín", "M269", "CAMEL", 2],
                   ["", "Comodín", "TX110", "PINK", 10]]
    assert indu == [["Drago Tech, Soporte Drago Leonel", 6, "", "J23", "AQUA", 1],
                    ["", 8, "", "J23", "AQUA", 1]]
    wb = openpyxl.load_workbook(io.BytesIO(oe.generar_excel_odoo({"numero": 7, "items": items}, "Drago")))
    assert wb.sheetnames == ["Sin talle", "Indu"]
    h1 = [list(r) for r in wb["Sin talle"].iter_rows(values_only=True)]
    h2 = [list(r) for r in wb["Indu"].iter_rows(values_only=True)]
    assert h1 == [oe.CAB_SIN_TALLE, ["Drago", "Comodín", "M269", "CAMEL", 2], [None, "Comodín", "TX110", "PINK", 10]]
    assert h2 == [oe.CAB_INDU, ["Drago", 6, None, "J23", "AQUA", 1], [None, 8, None, "J23", "AQUA", 1]]
    assert oe.nombre_archivo({"numero": 7}) == "odoo_pedido_000007.xlsx"


def test_compra_rapida_excel_ida_y_vuelta():
    import compra_rapida as cr
    df = pd.DataFrame([
        {"sku": "M211_U_2059", "ean": "779", "producto_cod": "M211", "producto_nombre": "Mochila",
         "color": "AQUA", "talle": "U", "precio": 100.0},
        {"sku": "J23_6_1", "ean": "", "producto_cod": "J23", "producto_nombre": "Jean",
         "color": "AQUA", "talle": "6", "precio": 50.0},
    ])
    data = cr.excel_plantilla(df, {"M211_U_2059": 3})
    texto, err = cr.texto_desde_excel(data)
    assert err is None and texto == "M211_U_2059,3"
    assert cr.texto_desde_excel(b"no es excel")[1]


def test_excel_plantilla_v2():
    import io

    import openpyxl
    import compra_rapida as cr
    df = pd.DataFrame([
        {"sku": "A_U_1", "ean": "779", "producto_cod": "A", "producto_nombre": "Prod A",
         "color": "AQUA", "talle": "U", "precio": 100.0},
        {"sku": "B_4_2", "ean": "", "producto_cod": "B", "producto_nombre": "Prod B",
         "color": "PINK", "talle": "4", "precio": 50.0},
    ])
    cli = {"nombre_display": "Cliente Test", "lista_precios": 1, "descuento": 20.0}
    data = cr.excel_plantilla(df, {"A_U_1": 3}, cliente=cli, iva_pct=21.0,
                              links_foto={"A_U_1": "https://x/foto.jpg"})
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Compra rápida", "Pedido"]
    ws = wb["Compra rápida"]
    assert [c.value for c in ws[1]] == cr.COLS_PLANTILLA
    assert ws["I2"].value == "=G2*H2" and ws["H2"].value == 3
    assert ws["J2"].hyperlink.target == "https://x/foto.jpg"
    wp = wb["Pedido"]
    textos = [str(c.value) for fila in wp.iter_rows() for c in fila if c.value is not None]
    assert any("descuento cabecera 20%" in t for t in textos)
    assert any("IVA 21%" in t for t in textos)
    assert any("FILTER" in getattr(c.value, "text", str(c.value))
               for fila in wp.iter_rows() for c in fila if c.value is not None)
    # SKU y EAN ocultos (pero presentes: la reimportación los usa)
    assert ws.column_dimensions["A"].hidden and ws.column_dimensions["B"].hidden
    assert not ws.column_dimensions["C"].hidden
    # protegido: solo Cantidad editable
    assert ws.protection.sheet and wp.protection.sheet
    assert ws["H2"].protection.locked is False and ws["G2"].protection.locked
    # se reimporta tal cual (lee la hoja 1)
    assert cr.texto_desde_excel(data) == ("A_U_1,3", None)
    # con miniaturas, la columna A queda para la imagen y las fórmulas corren una col
    from PIL import Image
    b = io.BytesIO(); Image.new("RGB", (4, 4), "red").save(b, "JPEG")
    data2 = cr.excel_plantilla(df, {}, cliente=cli, miniaturas={"A_U_1": b.getvalue()})
    ws2 = openpyxl.load_workbook(io.BytesIO(data2))["Compra rápida"]
    assert ws2["B1"].value == "SKU" and ws2["J2"].value == "=H2*I2"


def test_texto_desde_excel_multihoja():
    import io

    import xlsxwriter
    import compra_rapida as cr
    b = io.BytesIO(); wb = xlsxwriter.Workbook(b, {"in_memory": True})
    s1 = wb.add_worksheet("Resumen"); s1.write_row(0, 0, ["nota"]); s1.write(1, 0, "hola")
    s2 = wb.add_worksheet("Datos"); s2.write_row(0, 0, ["SKU", "Cantidad"]); s2.write_row(1, 0, ["X_U_1", 2])
    wb.close()
    assert cr.texto_desde_excel(b.getvalue()) == ("X_U_1,2", None)


def test_items_modificados():
    import pedidos as pe
    items = [{"sku": "A", "precio_unit": 100.0, "cantidad": 3, "subtotal": 300.0},
             {"sku": "B", "precio_unit": 50.0, "cantidad": 2, "subtotal": 100.0}]
    out = pe.items_modificados(items, {"A": 1})          # baja A a 1, B queda
    assert [(i["sku"], i["cantidad"], i["subtotal"]) for i in out] == [("A", 1, 100.0), ("B", 2, 100.0)]
    out = pe.items_modificados(items, {"B": 0})          # 0 quita la línea
    assert [i["sku"] for i in out] == ["A"]
    with pytest.raises(ValueError):
        pe.items_modificados(items, {"A": 0, "B": 0})    # vacío → error
    assert items[0]["cantidad"] == 3                     # no muta el original


def test_template_modificado():
    import email_notif as en
    assert "modificado" in en.EVENTOS and "modificado" in en.DEFAULT_TEMPLATES
    assert "{detalle}" in en.DEFAULT_TEMPLATES["modificado"]["cuerpo"]


def test_resumen_cambios():
    import pedidos as pe
    orig = [{"sku": "A", "producto_cod": "M1", "producto_nombre": "Moch", "color": "AQUA",
             "talle": "U", "cantidad": 3},
            {"sku": "B", "producto_cod": "T2", "producto_nombre": "Rem", "color": "PINK",
             "talle": "4", "cantidad": 2}]
    nuevos = [{"sku": "A", "cantidad": 1}]
    txt = pe.resumen_cambios(orig, nuevos)
    assert "M1 Moch | AQUA | T U: 3 u → 1 u" in txt
    assert "QUITADO: T2 Rem | PINK | T 4 (eran 2 u)" in txt
    assert pe.resumen_cambios(orig, orig) == ""
    import email_notif as en
    assert "{cambios}" in en.DEFAULT_TEMPLATES["modificado"]["cuerpo"]
