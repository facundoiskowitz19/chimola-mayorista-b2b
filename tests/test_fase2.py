"""Tests fase 2: overrides (catálogo/cliente/config), compra rápida, repetir pedido."""
from __future__ import annotations

import os
import sys
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ.setdefault("JWT_KEY", "test-key-no-usar")

import catalog  # noqa: E402
import compra_rapida as cr  # noqa: E402
import overrides  # noqa: E402
import pedidos  # noqa: E402


def _df():
    rows = [
        ("M211", "Mochila Soft Rainbow", "Chimola", "2059", "AQUA", "U", "M211_U_2059", "779", 99, 32900.0, 10.0),
        ("M211", "Mochila Soft Rainbow", "Chimola", "2058", "RAINBOW", "U", "M211_U_2058", "780", 5, 32900.0, 10.0),
        ("VEST2", "Vestido Lino", "Lima", "2018", "GREEN", "8", "VEST2_8_2018", "", 3, 50000.0, 0.0),
    ]
    cols = ["producto_cod", "producto_nombre", "marca", "color_cod", "color", "talle", "sku", "ean",
            "stock", "precio1", "descvta"]
    df = pd.DataFrame(rows, columns=cols)
    for c in catalog.PRECIO_COLS:
        if c not in df:
            df[c] = 0.0
    df["temporada"] = "SS26"
    df["rubro"] = "X"
    df["subrubro"] = None
    df["descripcion"] = ""
    return df


def test_aplicar_overrides(monkeypatch):
    ov = {
        "M211": {"publicado": False},
        "VEST2": {"nombre": "Vestido Edit", "precios": {"1": 44000}, "destacado": True},
    }
    monkeypatch.setattr(overrides, "get_catalogo_overrides", lambda: ov)
    out = overrides.aplicar_overrides(_df())
    assert set(out["producto_cod"]) == {"VEST2"}                      # M211 oculto
    assert out.iloc[0]["producto_nombre"] == "Vestido Edit"           # nombre pisado
    assert out.iloc[0]["precio1"] == 44000 and out.iloc[0]["destacado"]
    todo = overrides.aplicar_overrides(_df(), incluir_ocultos=True)
    assert set(todo["producto_cod"]) == {"M211", "VEST2"}
    assert todo[todo.producto_cod == "M211"]["publicado"].iloc[0] is False
    # sin overrides → intacto
    monkeypatch.setattr(overrides, "get_catalogo_overrides", lambda: {})
    assert len(overrides.aplicar_overrides(_df())) == 3


def test_override_cliente(monkeypatch):
    monkeypatch.setattr(overrides, "get_clientes_overrides",
                        lambda: {2722: {"descuento_pct": 25.0, "lista_precios": 2}})
    c = overrides.aplicar_override_cliente({"cliente_cod": 2722, "descuento": 20.0, "lista_precios": 1})
    assert c["descuento"] == 25.0 and c["descuento_origen"] == "Override"
    assert c["lista_precios"] == 2 and c["lista_origen"] == "Override"
    c2 = overrides.aplicar_override_cliente({"cliente_cod": 2723, "descuento": 30.0, "lista_precios": 1})
    assert c2["descuento"] == 30.0 and c2["descuento_origen"] == "Aleph"


def test_con_precio_descvta():
    df = _df()
    sin = catalog.con_precio(df, 1, aplicar_descvta=False)
    con = catalog.con_precio(df, 1, aplicar_descvta=True)
    assert sin[sin.sku == "M211_U_2059"]["precio"].iloc[0] == 32900
    assert con[con.sku == "M211_U_2059"]["precio"].iloc[0] == pytest.approx(32900 * 0.9)
    assert con[con.sku == "VEST2_8_2018"]["precio"].iloc[0] == 50000   # descvta 0


def test_parsear_lineas():
    txt = "M211_U_2059,3\n779;2\nM211_U_2058\t4\nVEST2_8_2018 5\n\nXXX,abc\n"
    out = cr.parsear_lineas(txt)
    assert out == [("M211_U_2059", 3, 1), ("779", 2, 2), ("M211_U_2058", 4, 3),
                   ("VEST2_8_2018", 5, 4), ("XXX", -1, 6)]


def test_resolver_pegado():
    df = catalog.con_precio(_df(), 1, aplicar_descvta=False)
    items, inc = cr.resolver_pegado("m211_u_2059,3\n779,2\nNOEXISTE,1\nM211_U_2058,99\nXXX,abc", df)
    por_sku = {i["sku"]: i for i in items}
    assert por_sku["M211_U_2059"]["cantidad"] == 5          # SKU + EAN consolidados
    assert por_sku["M211_U_2058"]["cantidad"] == 5          # recortado al stock
    por_linea = {i["linea"]: i for i in inc}
    assert por_linea[1]["tipo"] == "ok" and por_linea[2]["tipo"] == "ok"
    assert por_linea[3]["tipo"] == "no_encontrada"
    assert por_linea[4]["tipo"] == "ajustada" and por_linea[4]["pedido"] == 99 and por_linea[4]["cargado"] == 5
    assert por_linea[5]["tipo"] == "ilegible"
    assert cr.resumen_incidencias(inc) == {"agregadas": 2, "ajustadas": 1, "sin_reconocer": 2}


def test_repetir_pedido():
    df = catalog.con_precio(_df(), 1, aplicar_descvta=False)
    viejo = {"items": [
        {"sku": "M211_U_2059", "producto_nombre": "x", "color": "AQUA", "talle": "U", "cantidad": 3},
        {"sku": "M211_U_2058", "producto_nombre": "x", "color": "RAINBOW", "talle": "U", "cantidad": 10},
        {"sku": "YA_NO_EXISTE", "producto_nombre": "x", "color": "-", "talle": "U", "cantidad": 1},
    ]}
    items, avisos = pedidos.repetir_pedido(viejo, df)
    assert {i["sku"]: i["cantidad"] for i in items} == {"M211_U_2059": 3, "M211_U_2058": 5}
    assert items[0]["precio_unit"] == 32900                  # precio ACTUAL
    assert any("YA_NO_EXISTE" in a for a in avisos) and any("solo quedan 5" in a for a in avisos)


def test_estados_validos():
    assert pedidos.ESTADOS_SIGUIENTES["confirmado"] == ["procesado", "cancelado"]
    assert pedidos.ESTADOS_SIGUIENTES["cancelado"] == []


def test_puede_cancelar():
    cli = {"rol": "cliente", "cliente_cod": 1026, "email": "f@k.com"}
    ped = {"estado": "confirmado", "cliente_cod": 1026, "usuario_email": "f@k.com"}
    assert pedidos.puede_cancelar(ped, cli)
    assert not pedidos.puede_cancelar({**ped, "estado": "procesado"}, cli)      # ya lo tomó Lautin
    assert not pedidos.puede_cancelar({**ped, "cliente_cod": 2722}, cli)        # pedido ajeno
    assert not pedidos.puede_cancelar(ped, {**cli, "rol": "admin", "cliente_cod": None})
    assert not pedidos.puede_cancelar(None, cli) and not pedidos.puede_cancelar(ped, None)


def test_metricas_cliente():
    import datetime as dt

    import admin_ui
    mk = lambda n, est, dia, total, unid, items: {  # noqa: E731
        "numero": n, "estado": est, "total": total, "unidades": unid, "fecha_str": f"{dia:02d}/08/2026",
        "confirmed_at": dt.datetime(2026, 8, dia, 12, tzinfo=dt.timezone.utc), "items": items}
    lista = [
        mk(1, "procesado", 10, 100.0, 2, [{"producto_cod": "A", "producto_nombre": "a", "cantidad": 2}]),
        mk(2, "confirmado", 20, 300.0, 4, [{"producto_cod": "A", "producto_nombre": "a", "cantidad": 1},
                                           {"producto_cod": "B", "producto_nombre": "b", "cantidad": 3}]),
        mk(3, "cancelado", 21, 999.0, 9, [{"producto_cod": "C", "producto_nombre": "c", "cantidad": 9}]),
    ]
    m = admin_ui._metricas_cliente(lista)
    assert m["pedidos"] == 2 and m["cancelados"] == 1 and m["sin_procesar"] == 1
    assert m["total"] == 400.0 and m["unidades"] == 6 and m["ticket"] == 200.0
    assert m["ultimo"] == "20/08/2026"
    assert m["top"][0] == ["A", "a", 3] and all(r[0] != "C" for r in m["top"])
    vacio = admin_ui._metricas_cliente([])
    assert vacio["pedidos"] == 0 and vacio["ticket"] == 0.0 and vacio["ultimo"] == "—"


def test_cuerpos_email_estado():
    import email_notif
    ped = {"numero": 7, "cliente_nombre": "Kinderland", "cliente_cod": 1026,
           "fecha_str": "24/08/2026", "unidades": 10, "total": 1000.0, "usuario_email": "f@k.com"}
    proc = email_notif.cuerpo_estado(ped, "procesado", "admin@lautin.com.ar")
    assert "procesó tu pedido" in proc
    canc_cli = email_notif.cuerpo_estado(ped, "cancelado", "f@k.com")
    assert "CANCELADO por el cliente" in canc_cli
    canc_adm = email_notif.cuerpo_estado(ped, "cancelado", "admin@lautin.com.ar")
    assert "CANCELADO por el equipo de Lautin" in canc_adm


# ---------------------------------------------------------------------------
# Fase 3: overrides por variante, U.B., IVA, KPIs
# ---------------------------------------------------------------------------
def test_overrides_por_variante(monkeypatch):
    ov = {"M211": {"ub": 6, "variantes": {
        "M211_U_2059": {"stock": 10, "precios": {"1": 29000}},
        "M211_U_2058": {"oculta": True},
    }}}
    monkeypatch.setattr(overrides, "get_catalogo_overrides", lambda: ov)
    out = overrides.aplicar_overrides(_df())
    m59 = out[out.sku == "M211_U_2059"].iloc[0]
    assert m59["stock"] == 10 and m59["stock_manual"]            # stock manual reemplaza
    assert m59["precio1"] == 29000                                # precio por variante pisa
    assert m59["ub"] == 6
    assert "M211_U_2058" not in set(out["sku"])                  # variante oculta excluida
    adm = overrides.aplicar_overrides(_df(), incluir_ocultos=True)
    assert adm[adm.sku == "M211_U_2058"]["var_oculta"].iloc[0]   # admin la ve marcada
    # stock manual 0 → excluida para clientes
    ov["M211"]["variantes"]["M211_U_2059"]["stock"] = 0
    assert "M211_U_2059" not in set(overrides.aplicar_overrides(_df())["sku"])


def test_stock_actual_respeta_overrides(monkeypatch):
    import stock as stock_mod

    monkeypatch.setattr(stock_mod.bq_client, "query",
                        lambda sql, params=None: pd.DataFrame([{"sku": "A", "stock": 100},
                                                               {"sku": "B", "stock": 100}]))
    monkeypatch.setattr(overrides, "get_catalogo_overrides",
                        lambda: {"P": {"variantes": {"A": {"stock": 7}, "B": {"oculta": True}}}})
    assert stock_mod.stock_actual(["A", "B", "C"]) == {"A": 7, "B": 0, "C": 0}


def test_totales_con_iva():
    items = [{"sku": "A", "cantidad": 4, "precio_unit": 32900.0}]
    tot = pedidos.calcular_totales(items, 25, iva_pct=21)
    assert tot["total"] == 4 * 32900 * 0.75
    assert tot["iva_monto"] == pytest.approx(tot["total"] * 0.21)
    assert tot["total_con_iva"] == pytest.approx(tot["total"] * 1.21)
    sin = pedidos.calcular_totales(items, 25)
    assert sin["iva_pct"] == 0 and sin["total_con_iva"] == sin["total"]


def test_excel_con_iva():
    import datetime as dt
    import io

    import openpyxl
    items = [{"sku": "M211_U_2059", "ean": "779", "producto_cod": "M211", "producto_nombre": "Mochila",
              "color_cod": "2059", "color": "AQUA", "talle": "U", "cantidad": 3, "precio_unit": 32900.0}]
    p = {"numero": 99, "cliente_cod": 2722, "cliente_nombre": "Test", "cliente_cuit": "30-1",
         "usuario_email": "x@y.com", "lista_precios": 1, "items": items, "estado": "confirmado",
         "observaciones": "", "confirmed_at": dt.datetime(2026, 8, 23, 15, tzinfo=dt.timezone.utc),
         "fecha_str": "23/08/2026"}
    p.update(pedidos.calcular_totales(items, 20, iva_pct=21))   # muta items (agrega subtotal)
    p["xlsx_filename"] = pedidos.nombre_archivo(p)
    data = pedidos.generar_excel(p)
    ws = openpyxl.load_workbook(io.BytesIO(data))["Resumen"]
    vals = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, 30)}
    assert vals["IVA 21%"] == pytest.approx(p["iva_monto"])
    assert vals["TOTAL c/IVA"] == pytest.approx(p["total_con_iva"])


def test_kpis():
    import datetime as dt

    import admin_ui
    ahora = dt.datetime(2026, 8, 23, 12, 0, tzinfo=dt.timezone.utc)
    mk = lambda n, est, dia, total, unid, items: {  # noqa: E731
        "numero": n, "estado": est, "cliente_cod": 1, "total": total, "unidades": unid,
        "confirmed_at": dt.datetime(2026, 8, dia, 12, tzinfo=dt.timezone.utc), "items": items}
    lista = [
        mk(1, "confirmado", 20, 100.0, 2, [{"producto_cod": "A", "producto_nombre": "a", "cantidad": 2}]),
        mk(2, "procesado", 21, 200.0, 3, [{"producto_cod": "A", "producto_nombre": "a", "cantidad": 3}]),
        mk(3, "cancelado", 21, 999.0, 9, [{"producto_cod": "B", "producto_nombre": "b", "cantidad": 9}]),
        mk(4, "procesado", 1, 50.0, 1, [{"producto_cod": "C", "producto_nombre": "c", "cantidad": 1}]),
    ]
    k = admin_ui._kpis(lista, ahora)
    assert k["sin_procesar"] == 1
    assert k["pedidos_mes"] == 3 and k["monto_mes"] == 350.0 and k["unidades_mes"] == 6
    assert k["top"][0] == ["A", "a", 5]                    # cancelado excluido de montos, no de nada más
    assert all(row[0] != "B" for row in k["top"])


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-q"]))
